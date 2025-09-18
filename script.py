#!/usr/bin/env python3
import csv
import os
from openpyxl import load_workbook

# Define file paths
input_file = os.path.expanduser("~/downloads/thing.xlsx")
output_file = os.path.expanduser("~/downloads/thing.csv")

# Load the workbook and select active sheet
wb = load_workbook(input_file)
ws = wb.active

# Write to CSV
with open(output_file, 'w', newline='') as f:
    writer = csv.writer(f)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)

print(f"Converted: {output_file}")